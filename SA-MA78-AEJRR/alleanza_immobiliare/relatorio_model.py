import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from database import conectar

def exportar_corretores_excel():
    try:
        conexao = conectar()
        query = "SELECT id_corretor, nome, creci, telefone, email FROM Corretor"
        df = pd.read_sql(query, conexao)
        
        df.columns = ['ID', 'Nome Completo', 'CRECI', 'Telefone', 'E-mail']
        filename = "Relatorio_Corretores.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Corretores', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Corretores']
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            data_font = Font(name="Arial", size=10)
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
            )
            
            for col_idx, col in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left" if col_idx > 1 else "center", vertical="center")
            
            for row_idx in range(2, worksheet.max_row + 1):
                is_even = row_idx % 2 == 0
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border
                    if is_even:
                        cell.fill = zebra_fill
                    cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")
            
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        print(f"\n✅ Planilha executiva '{filename}' gerada com sucesso!")
    except Exception as e:
        print(f"\n❌ Erro ao gerar planilha: {e}")
    finally:
        conexao.close()


def exportar_corretores_pdf():
    try:
        conexao = conectar()
        query = "SELECT id_corretor, nome, creci, email FROM Corretor"
        
        # Oculta temporariamente os avisos do pandas no terminal
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_sql(query, conexao)
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # Cabeçalho do Relatório
        pdf.set_font("Arial", style="B", size=20)
        pdf.set_text_color(33, 78, 120) 
        pdf.cell(190, 10, txt="Alleanza Immobiliare", ln=True, align="L")
        
        pdf.set_font("Arial", size=10)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(190, 5, txt="Relatorio Executivo de Corretores Cadastrados", ln=True, align="L")
        pdf.ln(5)
        
        # Linha divisória azul (Corrigido para set_line_width)
        pdf.set_draw_color(33, 78, 120)
        pdf.set_line_width(0.5)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(10)
        
        # Cabeçalho da Tabela
        pdf.set_font("Arial", style="B", size=10)
        pdf.set_text_color(255, 255, 255) 
        pdf.set_fill_color(33, 78, 120)   
        
        pdf.cell(15, 8, "ID", border=1, align="C", fill=True)
        pdf.cell(65, 8, "Nome Completo", border=1, align="L", fill=True)
        pdf.cell(40, 8, "Inscricao CRECI", border=1, align="L", fill=True)
        pdf.cell(70, 8, "E-mail de Contato", border=1, align="L", fill=True)
        pdf.ln()
        
        # Dados da Tabela
        pdf.set_font("Arial", size=9)
        pdf.set_text_color(0, 0, 0)
        
        for idx, row in df.iterrows():
            fill = idx % 2 == 0
            pdf.set_fill_color(242, 245, 248) if fill else pdf.set_fill_color(255, 255, 255)
            
            pdf.cell(15, 7, str(row['id_corretor']), border=1, align="C", fill=True)
            pdf.cell(65, 7, str(row['nome']), border=1, align="L", fill=True)
            pdf.cell(40, 7, str(row['creci']), border=1, align="L", fill=True)
            pdf.cell(70, 7, str(row['email']), border=1, align="L", fill=True)
            pdf.ln()
            
        pdf.output("Relatorio_Corretores.pdf")
        print("\n✅ Documento PDF Executivo 'Relatorio_Corretores.pdf' gerado com sucesso!")
        
    except Exception as e:
        print(f"\n❌ Falha ao compilar PDF: {e}")
    finally:
        conexao.close()